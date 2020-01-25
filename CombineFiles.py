# Reads all the files from folder PythonFiles/ReadFiles
# Reads data in the first column of all the files in the above folder
# creates a new file called ToTalSample in the same folder
# NEw file contains aggregated first column from all the files.

import os
import openpyxl
listnum = []
FileList = os.listdir('./ReadFiles')
for files in FileList:
    FilePath = os.path.join('./ReadFiles',files)
    wb = openpyxl.load_workbook(FilePath)
    sheet = wb['Sheet']

    for i in range(1,sheet.max_row + 1):
        listnum.append(sheet.cell(i,1).value)

wbNew = openpyxl.Workbook()
sheet = wbNew.active

for i in range(0,len(listnum)):
    sheet.cell(i+1,1).value = listnum[i]


FilePath = os.path.join('./ReadFiles','TotalSample.xlsx')
wbNew.save(FilePath)

    



        
        
    
    
    
